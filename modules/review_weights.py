import numpy as npy
import openpyxl as xls
from modules.woopsies import Woopsies as woops

def single_contact(xlsheet, ipn, contact_nr, side, woopsy):
    temp_list_1 = [xlsheet.cell(3+ 144*side + 36*contact_nr +9*0+k,ipn).value for k in range(9)]
    rigidity_var = npy.array([float(val) if (not (val in [None, '-'])) else npy.nan for val in temp_list_1], dtype='float32')
    temp_list_2 = [xlsheet.cell(3+ 144*side + 36*contact_nr +9*1+k,ipn).value for k in range(9)]
    akinesia_var = npy.array([float(val) if (not (val in [None, '-'])) else npy.nan for val in temp_list_2], dtype='float32')
    temp_list_3 = [xlsheet.cell(3+ 144*side + 36*contact_nr +9*2+k,ipn).value for k in range(9)]
    tremor_var = npy.array([float(val) if (not (val in [None, '-'])) else npy.nan for val in temp_list_3], dtype='float32')

    temp_list_4 = [xlsheet.cell(3+ 144*side + 36*contact_nr +9*3+k,ipn).value for k in range(9)]
    side_effects_var = npy.array([False if val in [None, '-', '0',0] else True for val in temp_list_4], dtype='bool')
    side_effects_min = npy.zeros_like(side_effects_var,dtype='bool')

    for element in range(1,len(side_effects_var)):
        side_effects_min[element-1] = npy.max(side_effects_var[:element])

    side_effects_min[-1] = npy.max(side_effects_var)

    if npy.isnan(rigidity_var).all():
        rigidity_var = npy.zeros_like(rigidity_var)
    if npy.isnan(akinesia_var).all():
        akinesia_var = npy.zeros_like(akinesia_var)
    if npy.isnan(tremor_var).all():
        tremor_var = npy.zeros_like(tremor_var)

    temp_array = rigidity_var + akinesia_var + tremor_var
    
    for index in range(len(temp_array)):
        if side_effects_min[index] == True:
            temp_array[index] = npy.NaN
    woopsy.add_info('review_weights', f'calculating single contact side: {side}, contact: {contact_nr}')

    if temp_array[0] == 0:
        woopsy.add_woopsie('review_weights', f'clinical review starts with 0 score for {woopsy.get_subID()}, side: {side}, contact: {contact_nr}')
        weighted_improvement = npy.zeros([npy.size(temp_array)-1], dtype='float32')
    else:
        woopsy.add_info('review_weights', f'starting score ok for side: {side}, contact: {contact_nr}')
        if len(temp_array) > 1:
            bounded_diff = npy.diff(temp_array).copy() ### set upper bound to 0, if improvement is negative
            for i in range(npy.size(bounded_diff)):
                if bounded_diff[i] > 0:
                    bounded_diff[i] = 0
            # calculate scaling vector for linear scaling
            max_n_steps = 8
            actual_n_steps = len(bounded_diff)
            full_scaling_vector = (9-npy.arange(1, max_n_steps+1))/8
            scaling_vector = full_scaling_vector[:actual_n_steps]

            weighted_improvement = npy.cumsum(bounded_diff * scaling_vector) / temp_array[0]
            woopsy.add_info('review_weights', f'calculated contact {woopsy.get_subID()}, side: {side}, contact: {contact_nr}')
        else:
            woopsy.add_woopsie('review_weights', f'no valid clinical review after 1st step for {woopsy.get_subID()}, side: {side}, contact: {contact_nr}')
            weighhted_improvement = npy.zeros([npy.size(temp_array)-1], dtype='float32')
    return weighted_improvement


def calculate_weights(xlsheet, ipn_, side, woopsy):
    weighted_values = []

    for contact in range(4):
        woopsy.add_info('review_weights', f'calculating weights for {woopsy.get_subID()}, side: {side}, contact: {contact} -- 1st step')
        weights = single_contact(xlsheet,ipn_, contact, side, woopsy)
        if npy.isnan(weights).all():
            weighted_values.append(npy.zeros_like(weights, dtype='float32'))
            print(f'not enough valid clinical review weights for {woopsy.get_subID()}, side: {side}, contact: {contact}')
            print('setting to 1')
            woopsy.add_woopsie('review_weights', f'not enough valid clinical review weights for {woopsy.get_subID()}, side: {side}, contact: {contact} -- setting to 1')
        else:
            weighted_values.append(single_contact(xlsheet,ipn_, contact, side, woopsy))
            woopsy.add_info('review_weights', f'calculating weights for {woopsy.get_subID()}, side: {side}, contact: {contact} -- 2nd step')
        
    try:
        woopsy.add_info('review_weights', f'normalizing weights for {woopsy.get_subID()}, side: {side} -- 3rd step')
        if npy.isnan(weighted_values).all():
            not_normalized = [npy.array([0], dtype='float32'),npy.array([0], dtype='float32'),npy.array([0], dtype='float32'),npy.array([0], dtype='float32')]
            normalized = not_normalized
            woopsy.add_woopsie('review_weights', f'not enough valid clinical review weights for {woopsy.get_subID()}, side: {side} -- not enough valid weights')
        else:
            if npy.nanmean(npy.asarray(weighted_values)) != 0:
                not_normalized = [npy.nanmin(npy.asarray(weighted_values[contact])) / npy.nanmean(npy.asarray(weighted_values)) for contact in range(4)]
                normalized = [val / npy.sum(not_normalized) for val in not_normalized]
                woopsy.add_info('review_weights', f'normalized weights for {woopsy.get_subID()}, side: {side} -- 4th step')
            else:
                normalized = npy.ones(4, dtype='float32') / 4.0
                woopsy.add_info('review_weights', f'missing points -- setting weights for {woopsy.get_subID()}, side: {side} to 1')

        expanded_normalized = npy.zeros(8, dtype='float32')

        expanded_normalized[0] = normalized[0]
        expanded_normalized[1:4] = normalized[1]
        expanded_normalized[4:7] = normalized[2]
        expanded_normalized[7] = normalized[3]
    except:
        print(f'Error calculating weights for {woopsy.get_subID()}, side: {side}')
        print('setting to 1')
        woopsy.add_woopsie('review_weights', f'Error calculating weights for {woopsy.get_subID()}, side: {side} -- setting to 1')
        expanded_normalized = npy.ones(8, dtype='float32') / 4.0

    return expanded_normalized


def calculate_nudge(review_file, _sub_ID, woopsy):
    xlf = xls.open(review_file, read_only=True,keep_vba=False,data_only=True,keep_links=False)
    xl = xlf.active
    woopsy.add_info('review_weights', f'calculating nudge for sub-{_sub_ID} from {review_file}..')
    for row in xl.values:
        row_length = len(row)
        break

    ipn_first = 4 + 1
    ipn_last = row_length +1
    
    ipns_dict = dict({})
    for _ipn_ in range(ipn_first,ipn_last):
        ipn_label = 'sub-' + str.upper(xl.cell(1,_ipn_).value)
        ipns_dict[ipn_label] = _ipn_
    
    ipn = ipns_dict[_sub_ID]
    woopsy.add_info('review_weights', f'{_sub_ID} found.')

    nudge_L = calculate_weights(xl, ipn, 1, woopsy) / 0.25
    nudge_R = calculate_weights(xl, ipn, 0, woopsy) / 0.25
    woopsy.add_info('review_weights', f'nudge calculated for {_sub_ID}.')

    contact_nudge_R = 2 - nudge_R
    contact_nudge_L = 2 - nudge_L
    current_nudge_R = nudge_R
    current_nudge_L = nudge_L

    return contact_nudge_R, contact_nudge_L, current_nudge_R, current_nudge_L


if __name__ == '__main__':
    print('not called')